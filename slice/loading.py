from openpyxl import load_workbook
import os
import sys


project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
sys.path.append(project_root)


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'slice.settings')

import django
django.setup()
from manualtracking.models import FoodItemsdb  

#Change excel file path according to your path 
file_path = r'C:\Users\MUHAMMED SABIR\Desktop\mini\ABBREV.xlsx'


wb = load_workbook(file_path)
ws = wb.active


for row in ws.iter_rows(min_row=2,max_row=8,values_only=True):
    name,energy,protien,carbohydrate,fat,GmWt_1,GmWt_Desc1,GmWt_2,GmWt_Desc2 = row
   
    FoodItemsdb.objects.create(name=name,energy=energy,protien=protien,carbohydrate=carbohydrate,fat=fat,GmWt_1=GmWt_1,GmWt_Desc1=GmWt_Desc1,GmWt_2=GmWt_2,GmWt_Desc2=GmWt_Desc2)

print('Data imported successfully')
